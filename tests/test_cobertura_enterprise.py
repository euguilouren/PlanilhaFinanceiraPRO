"""
Testes de cobertura para módulos enterprise adicionados nas Fases 3-6.
Cobre: Conciliador.conciliar_aproximado, PrestadorContas, Verificador,
       AnalistaFinanceiro (comparativo_periodos, classificar_impostos_br).
"""

import pandas as pd

# ── Conciliador.conciliar_aproximado ─────────────────────────────


class TestConciliadorAproximado:
    def _df1(self):
        return pd.DataFrame(
            {
                "Valor": [100.0, 200.0, 300.0],
                "Data": ["01/01/2024", "05/01/2024", "10/01/2024"],
                "Entidade": ["Cliente A", "Cliente B", "Cliente C"],
            }
        )

    def _df2(self):
        return pd.DataFrame(
            {
                "Valor": [100.0, 205.0, 999.0],
                "Data": ["01/01/2024", "06/01/2024", "20/01/2024"],
                "Entidade": ["Cliente A", "Cliente B", "Outro"],
            }
        )

    def test_retorna_dataframe(self):
        from toolkit_financeiro import Conciliador

        result = Conciliador.conciliar_aproximado(self._df1(), self._df2(), "Valor", "Valor", "Data", "Data")
        assert isinstance(result, pd.DataFrame)
        assert not result.empty

    def test_colunas_esperadas(self):
        from toolkit_financeiro import Conciliador

        result = Conciliador.conciliar_aproximado(self._df1(), self._df2(), "Valor", "Valor")
        assert "Status" in result.columns
        assert "Score_Match" in result.columns
        assert "Diferença_R$" in result.columns

    def test_match_exato_recebe_score_alto(self):
        from toolkit_financeiro import Conciliador

        df1 = pd.DataFrame({"V": [100.0], "D": ["01/01/2024"]})
        df2 = pd.DataFrame({"V": [100.0], "D": ["01/01/2024"]})
        result = Conciliador.conciliar_aproximado(df1, df2, "V", "V", "D", "D")
        assert result["Score_Match"].iloc[0] >= 3

    def test_sem_match_score_zero(self):
        from toolkit_financeiro import Conciliador

        df1 = pd.DataFrame({"V": [100.0], "D": ["01/01/2024"]})
        df2 = pd.DataFrame({"V": [9999.0], "D": ["31/12/2024"]})
        result = Conciliador.conciliar_aproximado(df1, df2, "V", "V", "D", "D")
        assert result["Score_Match"].iloc[0] == 0

    def test_com_entidade_aumenta_score(self):
        from toolkit_financeiro import Conciliador

        df1 = pd.DataFrame({"V": [100.0], "D": ["01/01/2024"], "E": ["Empresa X"]})
        df2 = pd.DataFrame({"V": [100.0], "D": ["01/01/2024"], "E": ["Empresa X"]})
        result = Conciliador.conciliar_aproximado(df1, df2, "V", "V", "D", "D", "E", "E")
        assert result["Score_Match"].iloc[0] >= 5

    def test_itens_nao_encontrados_na_fonte2(self):
        from toolkit_financeiro import Conciliador

        df1 = pd.DataFrame({"V": [1.0]})
        df2 = pd.DataFrame({"V": [1.0, 2.0, 3.0]})
        result = Conciliador.conciliar_aproximado(df1, df2, "V", "V")
        assert len(result) >= 3  # 1 match + 2 não encontrados em Fonte_1


# ── PrestadorContas ───────────────────────────────────────────────


class TestPrestadorContas:
    def _df(self):
        return pd.DataFrame(
            {
                "Valor": [1000.0, -500.0, 200.0, -100.0],
                "Categoria": ["VENDAS", "CUSTO", "RECEITAS", "ALUGUEL"],
                "Tipo": ["RECEITA", "DESPESA", "RECEITA", "DESPESA"],
            }
        )

    def test_demonstrativo_movimentacao_com_tipo(self):
        from toolkit_financeiro import PrestadorContas

        result = PrestadorContas.demonstrativo_movimentacao(
            self._df(), "Valor", "Categoria", "Data", "Tipo", 500.0, "Jan/2024"
        )
        assert isinstance(result, pd.DataFrame)
        assert not result.empty
        assert "Tipo" in result.columns

    def test_demonstrativo_movimentacao_sem_tipo(self):
        from toolkit_financeiro import PrestadorContas

        result = PrestadorContas.demonstrativo_movimentacao(self._df(), "Valor", "Categoria", "Data")
        assert not result.empty
        assert "ENTRADAS" in result["Descrição"].values

    def test_demonstrativo_inclui_saldo_final(self):
        from toolkit_financeiro import PrestadorContas

        result = PrestadorContas.demonstrativo_movimentacao(self._df(), "Valor", "Categoria", "Data", None, 1000.0)
        saldo_final_row = result[result["Tipo"] == "SALDO"].iloc[-1]
        assert saldo_final_row["Valor"] is not None

    def test_orcado_vs_realizado(self):
        from toolkit_financeiro import PrestadorContas

        df_real = pd.DataFrame({"Cat": ["A", "B"], "Valor": [105.0, 190.0]})
        df_orc = pd.DataFrame({"Cat": ["A", "B"], "Valor": [100.0, 200.0]})
        result = PrestadorContas.orcado_vs_realizado(df_real, df_orc, "Cat", "Valor", "Valor")
        assert "Desvio_RS" in result.columns
        assert "Desvio_%" in result.columns
        assert "Status" in result.columns
        assert "Execução_%" in result.columns

    def test_orcado_vs_realizado_detecta_desvio_critico(self):
        from toolkit_financeiro import PrestadorContas

        df_real = pd.DataFrame({"Cat": ["A"], "Valor": [200.0]})
        df_orc = pd.DataFrame({"Cat": ["A"], "Valor": [100.0]})
        result = PrestadorContas.orcado_vs_realizado(df_real, df_orc, "Cat", "Valor", "Valor")
        assert result["Status"].iloc[0] == "DESVIO CRÍTICO — JUSTIFICAR"

    def test_resumo_saldos(self):
        from toolkit_financeiro import PrestadorContas

        contas = {
            "Caixa": {"saldo_inicial": 1000.0, "entradas": 500.0, "saidas": 200.0},
            "Banco": {"saldo_inicial": 5000.0, "entradas": 1000.0, "saidas": 2000.0},
        }
        result = PrestadorContas.resumo_saldos(contas, "Jan/2024")
        assert isinstance(result, pd.DataFrame)
        assert "Conta" in result.columns
        assert "Saldo_Final" in result.columns
        assert len(result) >= 3  # 2 contas + total

    def test_resumo_saldos_vazio(self):
        from toolkit_financeiro import PrestadorContas

        result = PrestadorContas.resumo_saldos({})
        assert isinstance(result, pd.DataFrame)


# ── AnalistaFinanceiro — métodos adicionais ───────────────────────


class TestAnalistaFinanceiroExtra:
    def _df(self):
        return pd.DataFrame(
            {
                "Data": ["01/01/2024", "01/02/2024", "01/03/2024", "01/01/2024"],
                "Valor": [1000.0, 1200.0, 900.0, 500.0],
                "Categoria": ["RECEITA", "RECEITA", "RECEITA", "DESPESA"],
            }
        )

    def test_comparativo_periodos_sem_categoria(self):
        from toolkit_financeiro import AnalistaFinanceiro

        result = AnalistaFinanceiro.comparativo_periodos(self._df(), "Valor", "Data", freq="M")
        assert isinstance(result, pd.DataFrame)
        assert not result.empty

    def test_comparativo_periodos_com_categoria(self):
        from toolkit_financeiro import AnalistaFinanceiro

        result = AnalistaFinanceiro.comparativo_periodos(self._df(), "Valor", "Data", "Categoria", freq="M")
        assert isinstance(result, pd.DataFrame)

    def test_classificar_impostos_br(self):
        from toolkit_financeiro import AnalistaFinanceiro

        df = pd.DataFrame(
            {
                "Categoria": [
                    "ICMS sobre vendas",
                    "IRPJ exercício",
                    "INSS folha",
                    "OUTRAS DESPESAS",
                ]
            }
        )
        result = AnalistaFinanceiro.classificar_impostos_br(df, "Categoria")
        assert "Classificação_DRE" in result.columns
        assert result["Classificação_DRE"].iloc[0] == "Dedução de Receita"
        assert result["Classificação_DRE"].iloc[1] == "IR/CSLL (após resultado)"
        assert result["Classificação_DRE"].iloc[2] == "Despesa Operacional (encargos)"
        assert result["Classificação_DRE"].iloc[3] == "Verificar classificação"


# ── Verificador ───────────────────────────────────────────────────


class TestVerificador:
    def test_verificar_atualizacao_ok(self):
        from toolkit_financeiro import Status, Verificador

        df_orig = pd.DataFrame({"NF": ["001"], "Valor": [100.0]})
        df_novos = pd.DataFrame({"NF": ["002"], "Valor": [50.0]})
        df_result = pd.DataFrame({"NF": ["001", "002"], "Valor": [100.0, 50.0]})
        result = Verificador.verificar_atualizacao(df_orig, df_novos, df_result, "Valor", ["NF"])
        assert result["status"] == Status.OK
        assert result["alertas"] == []

    def test_verificar_atualizacao_soma_divergente(self):
        from toolkit_financeiro import Verificador

        df_orig = pd.DataFrame({"NF": ["001"], "Valor": [100.0]})
        df_novos = pd.DataFrame({"NF": ["002"], "Valor": [50.0]})
        df_result = pd.DataFrame({"NF": ["001", "002"], "Valor": [100.0, 60.0]})
        result = Verificador.verificar_atualizacao(df_orig, df_novos, df_result, "Valor", ["NF"])
        assert result["status"] == "FALHA"
        tipos = [a["tipo"] for a in result["alertas"]]
        assert "SOMA_ATUALIZACAO_DIVERGENTE" in tipos

    def test_verificar_atualizacao_duplicatas_pos_update(self):
        from toolkit_financeiro import Verificador

        df_orig = pd.DataFrame({"NF": ["001"], "Valor": [100.0]})
        df_novos = pd.DataFrame({"NF": ["002"], "Valor": [50.0]})
        df_result = pd.DataFrame({"NF": ["001", "001"], "Valor": [100.0, 50.0]})
        result = Verificador.verificar_atualizacao(df_orig, df_novos, df_result, "Valor", ["NF"])
        tipos = [a["tipo"] for a in result["alertas"]]
        assert "DUPLICATAS_POS_ATUALIZACAO" in tipos

    def test_verificar_formulas_planilha(self, tmp_path):
        from openpyxl import Workbook

        from toolkit_financeiro import Verificador

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Valor"
        ws["A2"] = 100.0
        ws["A3"] = "TOTAL"
        ws["B3"] = 100.0  # valor fixo onde deveria ser =SUM()
        path = str(tmp_path / "test.xlsx")
        wb.save(path)
        result = Verificador.verificar_formulas_planilha(path)
        assert "abas_verificadas" in result
        assert "alertas" in result
        assert len(result["abas_verificadas"]) > 0

    def test_verificar_formulas_planilha_sem_alertas(self, tmp_path):
        from openpyxl import Workbook

        from toolkit_financeiro import Status, Verificador

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Valor"
        ws["A2"] = 100.0
        path = str(tmp_path / "limpo.xlsx")
        wb.save(path)
        result = Verificador.verificar_formulas_planilha(path)
        assert result["status"] == Status.OK

    def test_relatorio_verificacao_sem_alertas(self):
        from toolkit_financeiro import Status, Verificador

        verificacoes = [{"status": Status.OK, "descricao": "Verificação 1", "alertas": []}]
        relatorio = Verificador.relatorio_verificacao(verificacoes)
        assert "INTEGRIDADE CONFIRMADA" in relatorio
        assert "[OK]" in relatorio

    def test_relatorio_verificacao_com_alertas(self):
        from toolkit_financeiro import Status, Verificador

        verificacoes = [
            {
                "status": "FALHA",
                "descricao": "Verificação 2",
                "alertas": [{"severidade": Status.CRITICA, "mensagem": "Soma divergente"}],
            }
        ]
        relatorio = Verificador.relatorio_verificacao(verificacoes)
        assert "ALERTA" in relatorio
        assert "Soma divergente" in relatorio


# ── Verificador.verificar_integridade ────────────────────────────


class TestVerificadorIntegridade:
    def test_verificar_integridade_soma_ok(self):
        from toolkit_financeiro import Status, Verificador

        df_ent = pd.DataFrame({"Valor": [100.0, 200.0]})
        df_sai = pd.DataFrame({"Valor": [100.0, 200.0]})
        result = Verificador.verificar_integridade(df_ent, df_sai, col_valor="Valor")
        assert result["status"] == Status.OK

    def test_verificar_integridade_soma_divergente(self):
        from toolkit_financeiro import Verificador

        df_ent = pd.DataFrame({"Valor": [100.0, 200.0]})
        df_sai = pd.DataFrame({"Valor": [100.0, 250.0]})
        result = Verificador.verificar_integridade(df_ent, df_sai, col_valor="Valor")
        assert result["status"] == "FALHA"
